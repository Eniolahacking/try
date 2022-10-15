import openpyxl,os
you=openpyxl.Workbook()
ans=you.active()
okay=ans.cell(row=1,colunm=1)
okay.value="Budget"
two=ans.cell(row=1,colunm=1)
two.value="Name"
aa=ans['A2']
aa.value=30000000000
bb=ans["B2"]
bb.value="EFK"
you.save(r'\storage\emulated\0\try.xlsx')
print('Succes')
filo=r'\storage\emulated\0\try.xlsx'
yo=openpyxl.load_workbook(filo)
act=yo.active()
rowo=act.max_row()
print(rowo)
print("Name of row")
print("We will create a new sheet name Your choice")
tilo=int(input("input title:"))
yoooo=you.create_sheet(index=1,title=tilo)
yoooo.save(r'\storage\emulated\0\try.xlsx)
print("create successful")
po=yo.cell(row=1,colunm=1)
print(po.value)
os.starfile(r'/storage/emulated/0/Donwload/try.xlsx)
